# -*- coding: utf-8 -*-
"""
샘플 단일 엑셀 템플릿 생성기 (작성규칙 v1 적용)

출력: data/templates/샘플_2022_1회_v1.xlsx
  - 시트 1 「문항등록」  : v3 33컬럼 + 2행 헤더 + 2022년 1회 샘플 10문항
  - 시트 2 「작성규칙」  : 문항코드 규칙 + 허용/금지 태그 + LaTeX + inline 박스 + 이미지 파일명 규칙

내용은 data/templates/샘플_2022_1회_10문항.html 의 '원문(엑셀 입력)' 과 동일.
작성규칙대로: <ol> 미사용(①②③ 원문자) · 표는 LaTeX array · 박스는 inline style.
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- v3 컬럼 정의 (build_template_v3.py 와 동일) ----
COLS = [
    ('과정', 11, 'meta', '출처'), ('연도', 7, 'meta', '출처'), ('회차', 6, 'meta', '출처'),
    ('번호', 6, 'meta', '출처'), ('사용교재', 20, 'meta', '출처'), ('교재구분', 8, 'meta', '출처'),
    ('문항코드', 22, 'code', '코드'),
    ('강의주소', 26, 'video', '강의'),
    ('과목ID', 11, 'class', '분류'), ('챕터', 14, 'class', '분류'), ('대유형', 16, 'class', '분류'),
    ('중유형', 16, 'class', '분류'), ('내용', 22, 'class', '분류'),
    ('빈출도', 7, 'attr', '속성'), ('난이도', 7, 'attr', '속성'), ('문제유형', 10, 'attr', '속성'),
    ('변형이력', 8, 'attr', '속성'), ('비고', 16, 'attr', '속성'),
    ('발문', 45, 'stem', '발문'), ('조건', 18, 'stem', '발문'), ('발문그림', 14, 'img', '발문'),
    ('보기1', 22, 'opt', '보기'), ('보기1그림', 12, 'img', '보기'),
    ('보기2', 22, 'opt', '보기'), ('보기2그림', 12, 'img', '보기'),
    ('보기3', 22, 'opt', '보기'), ('보기3그림', 12, 'img', '보기'),
    ('보기4', 22, 'opt', '보기'), ('보기4그림', 12, 'img', '보기'),
    ('정답(1~4)', 8, 'ans', '정답'),
    ('해설', 50, 'exp', '해설'),
    ('오답분석', 40, 'wr', '오답분석'),
    ('학습포인트', 50, 'lp', '학습POINT'),
]

COLOR = {'meta':'475569','code':'4F46E5','video':'7C3AED','class':'0891B2','attr':'E11D48',
         'stem':'DB2777','opt':'2563EB','img':'D97706','ans':'0D9488','exp':'EA580C',
         'wr':'9333EA','lp':'059669'}
LITE = {'meta':'F1F5F9','code':'E0E7FF','video':'EDE9FE','class':'CFFAFE','attr':'FFE4E6',
        'stem':'FCE7F3','opt':'DBEAFE','img':'FEF3C7','ans':'CCFBF1','exp':'FFEDD5',
        'wr':'F3E8FF','lp':'D1FAE5'}

WHITE = Font(color='FFFFFF', bold=True, size=11)
HDR = Font(color='1F2937', bold=True, size=10)
CELL = Font(color='1F2937', size=10)
THIN = Side('thin', color='CCCCCC')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 박스 스니펫 (작성규칙 3-A / 3-B, inline style)
F = '<div style="border:1px solid #34d399;background:#f0fdf4;border-radius:8px;padding:10px 14px;margin:8px 0;">'
W = '<div style="border:1px solid #fbbf24;background:#fffbeb;border-radius:8px;padding:10px 14px;margin:8px 0;">'
E = '</div>'

# ---- 샘플 10문항 (HTML 원문과 동일, 규칙 적용) ----
def q(no, code3, subj, ch, big, mid, cont, bn, nd, pt, stem, o1,o2,o3,o4, ans, exp, wr='', lp=''):
    return {'과정':'전기기능사','연도':2022,'회차':1,'번호':no,
            '사용교재':'독끝 전기기능사 필기','교재구분':'A',
            '문항코드':f'elec_A_2022_01_{code3}','강의주소':'',
            '과목ID':subj,'챕터':ch,'대유형':big,'중유형':mid,'내용':cont,
            '빈출도':bn,'난이도':nd,'문제유형':pt,'변형이력':0,'비고':'',
            '발문':stem,'조건':'','발문그림':'',
            '보기1':o1,'보기1그림':'','보기2':o2,'보기2그림':'',
            '보기3':o3,'보기3그림':'','보기4':o4,'보기4그림':'',
            '정답(1~4)':ans,'해설':exp,'오답분석':wr,'학습포인트':lp}

SAMPLES = [
 q(1,'01','theory','정전기','콘덴서','정전용량','정전용량 비례 관계',5,2,'단답형',
   '콘덴서의 정전용량에 대한 설명으로 <strong>틀린</strong> 것은?',
   '전압에 반비례한다.','이동 전하량에 비례한다.','극판의 넓이에 비례한다.','극판의 간격에 비례한다.',4,
   "정전용량 $C$ 은 극판의 넓이 $A$ 에 비례하고 극판의 간격 $d$ 에 반비례한다. 따라서 ④ '극판의 간격에 비례한다'는 설명은 틀린 것이다.",
   "① $C=\\dfrac{Q}{V}$ 관계에 의해 전압에 반비례한다.<br>\n② $C=\\dfrac{Q}{V}$ 관계에 의해 전하량에 비례한다.<br>\n③ 정전용량 공식에 의해 넓이에 비례한다.",
   "정전용량 공식\n"+F+"\n$$C=\\dfrac{\\varepsilon S}{d}\\;[\\text{F}]$$\n$\\varepsilon$: 유전율[F/m] · $S$: 극판 단면적[m²] · $d$: 극판 간격[m]\n"+E),

 q(2,'02','theory','정전기','정전에너지','에너지 공식','정전에너지 식',5,2,'단답형',
   '정전에너지 $W[\\text{J}]$ 를 구하는 식으로 옳은 것은? (단, $C$ 는 콘덴서용량 $[\\text{F}]$, $V$ 는 공급전압 $[\\text{V}]$ 이다.)',
   '$W=\\dfrac{1}{2}CV^2$','$W=\\dfrac{1}{2}CV$','$W=\\dfrac{1}{2}C^2 V$','$W=2CV^2$',1,
   "정전에너지 공식\n"+F+"\n$$W=\\dfrac{1}{2}CV^2\\;[\\text{J}]$$\n$W$: 정전에너지[J] · $C$: 정전용량[F] · $V$: 전압[V]\n"+E),

 q(3,'05','theory','교류','3상 결선','Y결선','선간·상전압 관계',5,2,'단답형',
   '평형 3상 교류회로에서 Y결선할 때 선간전압 $V_l$ 과 상전압 $V_p$ 의 관계는?',
   '$V_l=V_p$','$V_l=\\sqrt{2}\\,V_p$','$V_l=\\sqrt{3}\\,V_p$','$V_l=\\dfrac{1}{\\sqrt{3}}V_p$',3,
   "Y 결선에서는 선간전압이 상전압의 $\\sqrt{3}$ 배이다.\n"+F+"\n$$V_l=\\sqrt{3}\\,V_p$$\n$V_l$: 선간전압[V] · $V_p$: 상전압[V]\n"+E),

 q(4,'07','theory','직류','전력','전력량과 전력','전력 계산',5,2,'계산형',
   '20분간 $876{,}000[\\text{J}]$ 의 일을 할 때 전력은 약 몇 $[\\text{kW}]$ 인가?',
   '0.73','7.3','73','730',1,
   "전력량이 $[\\text{J}]$ 단위이므로 시간을 초 $[\\text{s}]$ 로 환산해 공식에 대입한다.\n$$P=\\dfrac{876{,}000}{20\\times 60}=730\\,[\\text{W}]=0.73\\,[\\text{kW}]$$",
   '',
   "전력량과 전력\n"+F+"\n$$P=\\dfrac{W}{t}\\;[\\text{W}]$$\n$W$: 전력량[J] · $t$: 시간[s] · $P$: 전력[W]\n"+E),

 q(5,'11','theory','교류','순시값','주파수','각속도-주파수',5,3,'계산형',
   '다음 전압 파형의 주파수는 약 몇 $[\\text{Hz}]$ 인가?\n$$e=100\\sin\\!\\left(377t-\\dfrac{\\pi}{5}\\right)[\\text{V}]$$',
   '50','60','80','100',2,
   "주어진 식에서 각속도 $\\omega=2\\pi f=377$ 임을 알 수 있다. 따라서\n$$f=\\dfrac{377}{2\\pi}\\approx 60\\,[\\text{Hz}]$$",
   '',
   "교류 순시값 표현식\n"+F+"\n$$e=V_m\\sin(\\omega t+\\theta)=V_m\\sin(2\\pi f t+\\theta)\\;[\\text{V}]$$\n$V_m$: 전압 최댓값[V] · $\\omega$: 각속도[rad/s] · $\\theta$: 위상 · $f$: 주파수[Hz]\n"+E),

 q(6,'13','theory','정자계','쿨롱의 법칙','자기력 계산','두 자극 사이 힘',3,3,'계산형',
   '$m_1=4\\times10^{-5}[\\text{Wb}]$, $m_2=6\\times10^{-3}[\\text{Wb}]$, $r=10[\\text{cm}]$ 이면, 두 자극 $m_1,\\,m_2$ 사이에 작용하는 힘은 약 몇 $[\\text{N}]$ 인가?',
   '1.52','2.4','24','152',1,
   "길이를 $[\\text{m}]$ 단위로 환산해 쿨롱의 법칙에 대입한다.\n$$F=6.33\\times10^{4}\\times\\dfrac{4\\times10^{-5}\\times 6\\times10^{-3}}{(10\\times10^{-2})^2}\\approx 1.52\\,[\\text{N}]$$",
   '',
   "쿨롱의 법칙 (자기장)\n"+F+"\n$$F=k\\dfrac{m_1 m_2}{r^2}=6.33\\times10^{4}\\times\\dfrac{m_1 m_2}{r^2}\\;[\\text{N}]$$\n$k$: 쿨롱 상수 $(6.33\\times10^4)$ · $m_1,m_2$: 자극의 세기[Wb] · $r$: 거리[m]\n"+E),

 q(7,'23','machinery','동기기','전기자 반작용','반작용 구분','앞선 전류 반작용',3,3,'단답형',
   '동기 전동기 전기자 반작용에 대한 설명이다. 공급전압에 대한 <strong>앞선 전류</strong>의 전기자 반작용은?',
   '감자 작용','증자 작용','교차 자화 작용','편자 작용',1,
   "동기 전동기의 전기자 반작용은 부하 전류의 위상(역률)에 따라 다음과 같이 구분된다.\n"
   "$$\\begin{array}{c|c|c|c}\n\\text{전류 위상} & \\text{부하} & \\text{작용} & \\text{현상} \\\\ \\hline\n"
   "\\text{동상} & R & \\text{교차 자화} & \\text{주자속 찌그러짐} \\\\\n"
   "90^\\circ\\text{ 지상} & L & \\text{감자 작용} & \\text{기전력 감소} \\\\\n"
   "90^\\circ\\text{ 진상} & C & \\text{증자 작용} & \\text{기전력 증가}\n\\end{array}$$\n"
   "문제에서는 $90^\\circ$ 앞선(진상) 전류가 흐르므로 <strong>감자 작용</strong>을 하여 기전력을 감소시킨다.",
   '',
   "동기 발전기와 전동기는 $L,\\,C$ 부하에 대해 반작용 특성이 서로 반대로 발생한다.\n"
   +W+"\n<strong style=\"color:#b45309;\">주의</strong><br>\n발전기의 지상(L)=감자, 전동기의 지상(L)=증자. 발전기 기준으로만 외우면 전동기에서 반대로 틀린다.\n"+E),

 q(8,'32','machinery','변압기','효율','전부하 효율','효율 계산',3,3,'계산형',
   '$200[\\text{kVA}]$ 의 단상 변압기가 있다. 철손이 $1.6[\\text{kW}]$, 전부하 동손이 $2.4[\\text{kW}]$ 이다. 변압기의 역률이 $0.8$ 일 때 전부하 시의 효율 $[\\%]$ 은 약 얼마인가?',
   '96.6','97.6','98.6','99.6',2,
   "출력 $=200\\times0.8$, 손실 $=$ 철손 $+$ 동손 으로 두고 효율 공식에 대입한다.\n$$\\eta=\\dfrac{200\\times0.8}{200\\times0.8+1.6+2.4}\\times100\\approx 97.6\\,[\\%]$$",
   '',
   "변압기의 효율\n"+F+"\n$$\\eta=\\dfrac{P}{P+P_c+P_i}\\times100\\;[\\%]$$\n$P$: 출력 · $P_c$: 동손 · $P_i$: 철손\n"+E),

 q(9,'39','facility','반도체 소자','사이리스터','TRIAC','역병렬 SCR 등가 소자',5,2,'단답형',
   '역병렬 결합의 SCR의 특성과 같은 반도체 소자는?',
   'PUT','UJT','Diac','Triac',4,
   "역병렬 결합의 SCR과 동일한 특성을 가지며, 쌍방향성 3단자 사이리스터라고 불리는 소자는 <strong>TRIAC(트라이액)</strong>이다.",
   "① ② 발진 회로 등에 사용되는 소자이다.<br>\n③ 양방향성 2단자 소자로, 주로 TRIAC의 트리거용으로 사용된다.",
   'TRIAC의 심볼<br>\n<img src="elec_A_2022_01_39_lp1.png">'),

 q(10,'60','facility','옥내배선','특수장소','가연성 분진','분진 장소 배선 공사',3,2,'단답형',
   '소맥분, 전분 기타 가연성의 분진이 존재하는 곳의 저압 옥내 배선 공사 방법에 해당되는 것으로 짝지어진 것은?',
   '케이블 공사, 애자 사용 공사','금속관 공사, 콤바인 덕트관, 애자 사용 공사',
   '케이블 공사, 금속관 공사, 애자 사용 공사','케이블 공사, 금속관 공사, 합성수지관 공사',4,
   "가연성 분진(밀가루·전분 등)이 있는 장소는 폭발·화재 위험이 크므로 가장 튼튼하고 밀폐성이 좋은 공사 방법을 써야 한다. 밀폐성이 좋은 공사는 <strong>금속관 공사·합성수지관 공사·케이블 공사</strong>이므로 정답은 ④이다."),
]

# ---- 작성규칙 시트 ----
RULES = [
 ('CBT 문항 콘텐츠 작성 규칙 (v1)', ''),
 ('', ''),
 ('[문항코드 규칙]', ''),
 ('형식', 'elec_<교재구분>_<연도>_<회차2자리>_<번호2자리>'),
 ('예시', 'elec_A_2022_01_01  (교재 A · 2022년 · 1회 · 1번)'),
 ('비고', '정렬·식별 키.'),
 ('', ''),
 ('[자동 생성 — 문항코드]', ''),
 ('샘플(3~12행)', '문항코드 = 정적값(고정). 그대로 import 안전.'),
 ('작성행(13행~)', '문항코드 = 수식(연노랑). 출처만 채우면 즉시 표시 — 작성 편의용.'),
 ('수식', '=IF(COUNTA(...)=4,"elec_"&교재구분&"_"&연도&"_"&TEXT(회차,"00")&"_"&TEXT(번호,"00"),"")'),
 ('안전장치', '엑셀 수식은 저장 전 미계산일 수 있으나, 서버(일괄등록)가 출처로 문항코드를 재계산하므로 비어도 안전.'),
 ('이미지명', '엑셀에서 자동 생성하지 않음. 그림 업로드 시 서버가 {문항코드}_{슬롯}{순번}.png 로 부여(고정). 본문/그림셀엔 그 이름을 그대로 참조.'),
 ('', ''),
 ('[이미지 파일명 규칙]', ''),
 ('형식', '{문항코드}_{슬롯}{순번}.png  (업로드 시 자동 부여)'),
 ('슬롯', '발문=stem · 보기=opt1~opt4 · 해설=exp · 오답분석=wrong · 학습포인트=lp'),
 ('예시', 'elec_A_2022_01_39_lp1.png  (39번 학습포인트 1번 그림)'),
 ('참조', '셀에는 <img src="elec_A_2022_01_39_lp1.png"> 형태로 본문 어디에나'),
 ('', ''),
 ('[허용 HTML 태그]', ''),
 ('강조/첨자', '<strong> <em> <u> <sub> <sup>'),
 ('줄바꿈', '<br>'),
 ('그림', '<img src="파일명">'),
 ('박스(틀)', '<div style="..."> — 아래 표준 스니펫만'),
 ('특수문자', '&lt;는 < · &gt;는 > · &amp;는 &'),
 ('', ''),
 ('[금지 — 대체 방법]', ''),
 ('<ol><ul><li>', '줄마다 <br>, 번호는 원문자 ①②③④ 로 직접 입력'),
 ('<table>', '표·행렬은 LaTeX  $$\\begin{array}{...}\\end{array}$$'),
 ('<p>', '문단 구분은 <br><br>'),
 ('class="..."', '모양은 class 금지 → inline style(아래)만'),
 ('', ''),
 ('[수식 — LaTeX / KaTeX]', ''),
 ('인라인', '$ ... $   예) 저항 $R=\\dfrac{V}{I}$'),
 ('디스플레이', '$$ ... $$   예) $$P=VI=I^2 R$$'),
 ('가능 환경', 'array · matrix · cases · aligned'),
 ('불가 환경', 'enumerate · itemize · tabular  (목록은 <br>+원문자)'),
 ('', ''),
 ('[박스 표준 스니펫 — inline style]', ''),
 ('핵심공식(초록)', F + ' $$...$$  ... ' + E),
 ('주의(호박)', W + ' <strong style="color:#b45309;">주의</strong><br> ... ' + E),
 ('규칙', 'style 값 변경·박스 중첩 금지. 독끝 바깥 박스는 앱이 자동(셀에 안 씀).'),
 ('', ''),
 ('[제출 전 체크]', ''),
 ('1', '<ol> <ul> <li> <table> <p> 안 썼나?'),
 ('2', 'class= 안 썼나? 모양은 전부 style= 인가?'),
 ('3', '번호는 원문자 ①②③ 인가?'),
 ('4', '표는 <table> 아니라 $$\\begin{array}$$ 인가?'),
 ('5', '수식은 $…$ / $$…$$ 안에 있나?'),
 ('6', '이미지 파일명이 {문항코드}_{슬롯}{순번}.png 인가?'),
]


def build(out_path):
    wb = openpyxl.Workbook()

    # ===== 시트 1: 문항등록 =====
    ws = wb.active
    ws.title = '문항등록'
    n = len(COLS)

    # 1행: 그룹 헤더 (연속 같은 그룹 병합)
    i = 0
    while i < n:
        g = COLS[i][3]
        j = i
        while j + 1 < n and COLS[j + 1][3] == g:
            j += 1
        c = ws.cell(1, i + 1, g)
        if j > i:
            ws.merge_cells(start_row=1, start_column=i + 1, end_row=1, end_column=j + 1)
        c.fill = PatternFill('solid', fgColor=COLOR[COLS[i][2]])
        c.font = WHITE
        c.alignment = CTR
        c.border = BOX
        i = j + 1

    # 2행: 컬럼 헤더
    for idx, (name, w, tone, grp) in enumerate(COLS, 1):
        cc = ws.cell(2, idx, name)
        cc.fill = PatternFill('solid', fgColor=LITE[tone])
        cc.font = HDR
        cc.alignment = CTR
        cc.border = BOX
        ws.column_dimensions[get_column_letter(idx)].width = w

    # 컬럼명 → 엑셀 열문자
    L = {name: get_column_letter(idx) for idx, (name, *_ ) in enumerate(COLS, 1)}
    AUTO_FILL = PatternFill('solid', fgColor='FEF9C3')  # 연노랑 = 자동수식

    def code_formula(r):
        # 출처(교재구분 F·연도 B·회차 C·번호 D) 모두 채워지면 문항코드 자동
        return (f'=IF(COUNTA(${L["교재구분"]}{r},${L["연도"]}{r},${L["회차"]}{r},${L["번호"]}{r})=4,'
                f'"elec_"&${L["교재구분"]}{r}&"_"&${L["연도"]}{r}&"_"&'
                f'TEXT(${L["회차"]}{r},"00")&"_"&TEXT(${L["번호"]}{r},"00"),"")')

    # 3행~: 샘플 10문항 — 문항코드는 정적값(import 안전), 이미지 슬롯 비움
    for r, smp in enumerate(SAMPLES, 3):
        for idx, (name, w, tone, grp) in enumerate(COLS, 1):
            cell = ws.cell(r, idx, smp.get(name, ''))
            cell.font = CELL
            cell.border = BOX
            cell.alignment = CTR if (len(str(cell.value)) <= 8 and name != '발문') else WRAP
        ws.row_dimensions[r].height = 150

    # 13행~: 빈 작성용 행 (수식 미리 박음) — 출처 채우면 코드·이미지명 자동 형성
    base = 3 + len(SAMPLES)
    DEMO = {'과정': '전기기능사', '연도': 2022, '회차': 1, '번호': 11,
            '사용교재': '독끝 전기기능사 필기', '교재구분': 'A',
            '비고': '← 출처만 채운 데모행: 문항코드 자동 생성됨'}
    for k in range(5):
        r = base + k
        for idx, (name, w, tone, grp) in enumerate(COLS, 1):
            if name == '문항코드':
                cell = ws.cell(r, idx, code_formula(r)); cell.fill = AUTO_FILL
            else:
                cell = ws.cell(r, idx, DEMO.get(name, '') if k == 0 else '')
            cell.font = CELL
            cell.border = BOX
            cell.alignment = WRAP if name in ('발문', '비고') else CTR
        ws.row_dimensions[r].height = 26

    ws.freeze_panes = 'G3'  # 출처~코드 좌측, 헤더 2행 고정

    # ===== 시트 2: 작성규칙 =====
    gs = wb.create_sheet('작성규칙')
    gs.column_dimensions['A'].width = 22
    gs.column_dimensions['B'].width = 110
    for r, (a, b) in enumerate(RULES, 1):
        ca, cb = gs.cell(r, 1, a), gs.cell(r, 2, b)
        if a.startswith('[') or (a and not b and r == 1):
            ca.font = Font(bold=True, color='FFFFFF', size=11)
            ca.fill = PatternFill('solid', fgColor='334155')
            cb.fill = PatternFill('solid', fgColor='334155')
        else:
            ca.font = Font(bold=True, color='1F2937', size=10)
            cb.font = Font(color='1F2937', size=10)
        ca.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cb.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    gs.freeze_panes = 'A2'

    wb.save(out_path)
    print('saved:', out_path)


if __name__ == '__main__':
    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(here, '..', 'data', 'templates', '샘플_2022_1회_v1.xlsx')
    build(os.path.normpath(out))
