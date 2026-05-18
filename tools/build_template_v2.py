"""
v2 양식 생성:
  - data/templates/question-template-v2.xlsx (단순 콘텐츠 등록용)
  - data/templates/[전기기능사 필기]CBT_문항관리_통합템플릿_v2.xlsx (메타+콘텐츠 통합)
컬럼은 모두 HTML 허용. 그림은 별도 컬럼 (파일명).
PNG 샘플 3종(35번/41번/02번)을 검증용 행으로 채움.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from copy import copy

# ---------- 공통 스타일 ----------
HDR_FILL = PatternFill('solid', fgColor='FFE4EC')   # 연핑크
META_FILL = PatternFill('solid', fgColor='E8F0FE')  # 연파랑
CONTENT_FILL = PatternFill('solid', fgColor='FFF7E0')  # 연노랑
IMG_FILL = PatternFill('solid', fgColor='E8F8E8')   # 연초록
HDR_FONT = Font(bold=True, size=11)
THIN = Side(border_style='thin', color='CCCCCC')
HDR_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row=1, fills=None, freeze=True):
    """fills: dict {col_idx: PatternFill} - 없으면 기본 HDR_FILL"""
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = (fills.get(c, HDR_FILL) if fills else HDR_FILL)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = HDR_BORDER
    ws.row_dimensions[row].height = 32
    if freeze:
        ws.freeze_panes = 'A2'

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------- 콘텐츠 컬럼 (양쪽 공통) ----------
# (헤더명, 너비, 카테고리)  카테고리: meta/text/img/ans
CONTENT_COLS = [
    ('발문',          50, 'text'),
    ('조건',          25, 'text'),
    ('발문그림',      18, 'img'),
    ('보기1',         28, 'text'),
    ('보기1그림',     16, 'img'),
    ('보기2',         28, 'text'),
    ('보기2그림',     16, 'img'),
    ('보기3',         28, 'text'),
    ('보기3그림',     16, 'img'),
    ('보기4',         28, 'text'),
    ('보기4그림',     16, 'img'),
    ('정답(1~4)',     10, 'ans'),
    ('해설',          60, 'text'),
    ('해설그림',      18, 'img'),
    ('오답분석',      55, 'text'),
    ('학습포인트',    60, 'text'),
    ('학습포인트그림',18, 'img'),
]

# ---------- 샘플 데이터 (PNG 3종 → HTML) ----------
SAMPLE_Q35 = {
    '발문': '그림과 같은 분상 기동형 단상 유도 전동기를 역회전시키기 위한 방법이 아닌 것은?',
    '조건': '',
    '발문그림': 'q35_stem.png',
    '보기1': '원심력스위치를 개로 또는 폐로 한다.',
    '보기1그림': '',
    '보기2': '기동권선이나 운전권선의 어느 한 권선의 단자접속을 반대로 한다.',
    '보기2그림': '',
    '보기3': '기동권선의 단자접속을 반대로 한다.',
    '보기3그림': '',
    '보기4': '운전권선의 단자접속을 반대로 한다.',
    '보기4그림': '',
    '정답(1~4)': 1,
    '해설': '원심력 스위치는 전동기가 정격 속도의 75~80%에 도달했을 때 기동권선을 회로에서 분리하는 역할만 수행하며, 회전 방향에는 무관하다.',
    '해설그림': '',
    '오답분석': '<ol><li>두 권선 중 하나만 선택하여 극성을 바꾸면 회전 자계의 방향이 바뀌어 역회전하게 되는 올바른 방법이다.</li><li>보조권선(기동권선)의 접속을 바꾸어 역회전시키는 방법 중 하나이다.</li><li>주권선(운전권선)의 접속을 바꾸어 역회전시키는 방법 중 하나이다.</li></ol>',
    '학습포인트': '<p><strong>전동기 역회전 방법</strong></p><table><thead><tr><th>전동기 종류</th><th>역회전 방법</th></tr></thead><tbody><tr><td>분상 기동형</td><td>주권선 또는 보조권선 중 한 접속을 반대로</td></tr><tr><td>3상 유도 전동기</td><td>3선 중 임의의 2선의 접속을 서로 바꿈</td></tr><tr><td>단상 직권 정류자형</td><td>브러시의 위치를 이동시킴</td></tr><tr><td>쉐이딩 코일형</td><td>구조적으로 역회전이 불가능함</td></tr></tbody></table>',
    '학습포인트그림': '',
}

SAMPLE_Q41 = {
    '발문': '고압 가공전선로의 지지물로 철탑을 사용하는 경우 경간은 몇 [m] 이하로 제한하는가?',
    '조건': '',
    '발문그림': '',
    '보기1': '150', '보기1그림': '',
    '보기2': '300', '보기2그림': '',
    '보기3': '500', '보기3그림': '',
    '보기4': '600', '보기4그림': '',
    '정답(1~4)': 4,
    '해설': '전선로의 지지물 종류에 따라 전선이 처지지 않고 안전하게 유지될 수 있는 지지물 간의 거리(경간)가 규정되어 있다. 철탑은 강도가 가장 높아 경간을 가장 길게 할 수 있다.',
    '해설그림': '',
    '오답분석': '',
    '학습포인트': '<p><strong>지지물 종류별 보안공사 경간 비교</strong></p><table><thead><tr><th>지지물의 종류</th><th>일반 경간</th><th>보안공사 경간</th></tr></thead><tbody><tr><td>A종 (목주, 철주, 철근 콘크리트주)</td><td>150[m]</td><td>100[m]</td></tr><tr><td>B종 (철주, 철근 콘크리트주)</td><td>250[m]</td><td>150[m]</td></tr><tr><td>철탑</td><td>600[m]</td><td>400[m]</td></tr></tbody></table>',
    '학습포인트그림': '',
}

SAMPLE_Q02 = {
    '발문': '반자성체 물질의 특색을 나타낸 것은? (단, μ<sub>s</sub>는 비투자율이다.)',
    '조건': '',
    '발문그림': '',
    '보기1': 'μ<sub>s</sub> &gt; 1', '보기1그림': '',
    '보기2': 'μ<sub>s</sub> ≫ 1',   '보기2그림': '',
    '보기3': 'μ<sub>s</sub> = 1',   '보기3그림': '',
    '보기4': 'μ<sub>s</sub> &lt; 1', '보기4그림': '',
    '정답(1~4)': 4,
    '해설': '<p><strong>비투자율에 따른 분류</strong></p><table><thead><tr><th>종류</th><th>비투자율(μ<sub>s</sub>)</th><th>특징</th></tr></thead><tbody><tr><td>강자성체</td><td>μ<sub>s</sub> ≫ 1</td><td>자석에 잘 붙음 (철, 니켈, 코발트)</td></tr><tr><td>상자성체</td><td>μ<sub>s</sub> &gt; 1</td><td>자석에 약하게 붙음 (알루미늄, 백금)</td></tr><tr><td>반자성체</td><td>μ<sub>s</sub> &lt; 1</td><td>자석에서 밀려남 (구리, 은, 비스무트)</td></tr></tbody></table><p>반자성체(역자성체)는 비투자율 μ<sub>s</sub>가 1보다 작은 물질을 말한다. 따라서 정답은 μ<sub>s</sub> &lt; 1 이다.</p>',
    '해설그림': '',
    '오답분석': '<ol><li>상자성체에 대한 설명이다.</li><li>강자성체에 대한 설명이다.</li><li>진공 또는 공기의 비투자율이다.</li></ol>',
    '학습포인트': '',
    '학습포인트그림': '',
}

# ---------- 사용가이드 (양쪽 공통) ----------
GUIDE_ROWS = [
    ('컬럼', '설명', '필수', '예시'),
    ('번호', '문항 번호 (회차 내 고유). 정수.', '필수', '1'),
    ('과목ID', '커리큘럼ID 시트의 과목ID 그대로. (theory/machinery/facility)', '필수', 'theory'),
    ('토픽ID', '커리큘럼ID 시트 참고. 영문 소문자/언더스코어.', '필수', 'electromagnetism'),
    ('서브토픽ID', '커리큘럼ID 시트 참고.', '권장', 'mutual_induction'),
    ('발문', '문제 본문. HTML 허용 (<br>, <sub>, <sup>, <strong>, <em>).', '필수', '...전류가 변화하면...'),
    ('조건', '문제에 주어진 조건. 없으면 비움.', '선택', 'P=100[Ω], Q=10[Ω]'),
    ('발문그림', '발문에 들어가는 그림 파일명만 입력 (public/questions/*.png 기준).', '선택', 'q35_stem.png'),
    ('보기1~4', '4지선다 본문. HTML 허용. 보기 자체가 그림형이면 비우고 보기N그림만 채움.', '필수', '<img>는 ‘보기N그림’으로 분리'),
    ('보기1그림~보기4그림', '보기가 그림형일 때 파일명. 텍스트형이면 비움.', '선택', 'opt_q35_1.png'),
    ('정답(1~4)', '정답 보기 번호. 1·2·3·4 중 하나.', '필수', '4'),
    ('해설', '해설 본문. HTML 허용. <table> 가능.', '필수', '<table><tr><th>...</th></tr></table>'),
    ('해설그림', '해설 다이어그램 파일명.', '선택', 'q35_exp.png'),
    ('오답분석', '오답 각각의 분석. <ol><li>...</li></ol> 권장.', '선택', '<ol><li>...</li></ol>'),
    ('학습포인트', '독끝 학습 Point. HTML 허용. <table> 가능.', '선택', '<p><strong>...</strong></p><table>...</table>'),
    ('학습포인트그림', '학습포인트 다이어그램 파일명.', '선택', 'q35_lp.png'),
    ('', '', '', ''),
    ('[HTML 사용 규칙]', '', '', ''),
    ('인라인 강조', '<strong>, <em>, <u>, <code>', '', ''),
    ('첨자/지수', '<sub>2</sub>, <sup>2</sup>', '', 'μ<sub>s</sub>, mm<sup>2</sup>'),
    ('줄바꿈', '<br>', '', ''),
    ('리스트', '<ul><li>...</li></ul> 또는 <ol><li>...</li></ol>', '', ''),
    ('표', '<table><thead><tr><th>헤더</th></tr></thead><tbody><tr><td>셀</td></tr></tbody></table>', '', ''),
    ('이미지', '본문 안 인라인이 필요할 때만 <img src="파일명">. 보통은 그림 컬럼 사용.', '', ''),
    ('특수문자', '&lt;는 <, &gt;는 >, &amp;는 &, &nbsp;는 공백', '', '부등호는 반드시 &lt; &gt;'),
]

def write_guide(wb):
    if '사용가이드' in wb.sheetnames:
        del wb['사용가이드']
    g = wb.create_sheet('사용가이드')
    for row in GUIDE_ROWS:
        g.append(row)
    style_header(g, row=1, freeze=False)
    set_widths(g, [22, 65, 8, 40])
    # 본문 wrap
    for r in range(2, g.max_row + 1):
        for c in range(1, 5):
            g.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')

# ---------- 커리큘럼ID 시트 ----------
CURRICULUM = [
    ('과목ID', '과목명', '토픽ID', '토픽명', '서브토픽ID', '서브토픽명'),
    ('theory', '전기이론', 'dc_circuit', '직류회로', 'ohm', '옴의 법칙'),
    ('theory', '전기이론', 'dc_circuit', '직류회로', 'kirchhoff', '키르히호프 법칙'),
    ('theory', '전기이론', 'dc_circuit', '직류회로', 'series_parallel', '직병렬 회로'),
    ('theory', '전기이론', 'electromagnetism', '정자계', 'mutual_induction', '상호유도'),
    ('theory', '전기이론', 'electromagnetism', '정자계', 'magnetic_material', '자성체와 자기회로'),
    ('theory', '전기이론', 'ac_circuit', '교류회로', 'sinusoidal', '정현파 교류'),
    ('theory', '전기이론', 'machine_basic', '전동기 기초', 'single_phase_im', '단상 유도전동기'),
    ('facility', '전기설비', 'overhead_line', '가공전선로', 'span', '경간'),
]

def write_curriculum(wb):
    if '커리큘럼ID' in wb.sheetnames:
        del wb['커리큘럼ID']
    c = wb.create_sheet('커리큘럼ID')
    for row in CURRICULUM:
        c.append(row)
    style_header(c, row=1, freeze=False)
    set_widths(c, [12, 14, 18, 14, 22, 22])

# ---------- 1) question-template-v2.xlsx ----------
def build_simple():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '문항등록'

    # 메타 4개 + 콘텐츠 17개 = 21
    meta_cols = [('번호', 6, 'meta'), ('과목ID', 12, 'meta'),
                 ('토픽ID', 18, 'meta'), ('서브토픽ID', 22, 'meta')]
    all_cols = meta_cols + CONTENT_COLS

    headers = [h for h, _, _ in all_cols]
    ws.append(headers)

    widths = [w for _, w, _ in all_cols]
    set_widths(ws, widths)

    # 컬럼별 색상 매핑
    fills = {}
    for i, (_, _, cat) in enumerate(all_cols, 1):
        if cat == 'meta':
            fills[i] = META_FILL
        elif cat == 'img':
            fills[i] = IMG_FILL
        elif cat == 'ans':
            fills[i] = PatternFill('solid', fgColor='FFE0E0')
        else:
            fills[i] = CONTENT_FILL
    style_header(ws, row=1, fills=fills)

    # 헤더 코멘트
    ws.cell(1, headers.index('발문') + 1).comment = Comment('HTML 허용', '시스템')
    ws.cell(1, headers.index('해설') + 1).comment = Comment('HTML 허용. <table> 가능.', '시스템')
    ws.cell(1, headers.index('오답분석') + 1).comment = Comment('<ol><li>...</li></ol> 권장', '시스템')
    ws.cell(1, headers.index('학습포인트') + 1).comment = Comment('HTML 허용. <table> 가능.', '시스템')

    # 샘플 행 3개
    samples = [
        (1, 'machinery', 'single_phase_im', 'single_phase_im', SAMPLE_Q35),
        (41, 'facility', 'overhead_line', 'span', SAMPLE_Q41),
        (2, 'theory', 'electromagnetism', 'magnetic_material', SAMPLE_Q02),
    ]
    for no, subj, tp, st, data in samples:
        row = [no, subj, tp, st] + [data[h] for h, _, _ in CONTENT_COLS]
        ws.append(row)

    # 본문 셀 wrap
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 120
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')

    write_guide(wb)
    write_curriculum(wb)

    out = 'data/templates/question-template-v2.xlsx'
    wb.save(out)
    print(f'WRITE: {out}  ({ws.max_column} cols, {ws.max_row} rows)')

# ---------- 2) 통합템플릿_v2 ----------
META_COLS_FULL = [
    ('문항종류', 10), ('문항번호', 10), ('문항코드', 12), ('변형이력', 8),
    ('사용교재', 22), ('교재구분', 9), ('초판출간일', 12), ('등록자', 10), ('등록일', 12),
    ('과정', 12), ('연도', 8), ('회차', 6), ('번호', 6),
    ('과목', 12), ('챕터', 16), ('대유형', 22), ('중유형', 22), ('내용', 28),
    ('빈출도', 7), ('난이도', 7), ('문제유형', 10), ('비고', 16),
    ('문제파일명', 24), ('정답파일명', 24), ('해설파일명', 24),
]

def build_full():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '문항관리'

    headers = [h for h, _ in META_COLS_FULL] + [h for h, _, _ in CONTENT_COLS]
    ws.append(headers)

    widths = [w for _, w in META_COLS_FULL] + [w for _, w, _ in CONTENT_COLS]
    set_widths(ws, widths)

    fills = {}
    meta_n = len(META_COLS_FULL)
    for i in range(1, meta_n + 1):
        fills[i] = META_FILL
    # 이미지등록 3컬럼 표시
    img_register_idx = [headers.index('문제파일명')+1, headers.index('정답파일명')+1, headers.index('해설파일명')+1]
    for i in img_register_idx:
        fills[i] = IMG_FILL
    for j, (_, _, cat) in enumerate(CONTENT_COLS):
        col = meta_n + j + 1
        if cat == 'img':
            fills[col] = IMG_FILL
        elif cat == 'ans':
            fills[col] = PatternFill('solid', fgColor='FFE0E0')
        else:
            fills[col] = CONTENT_FILL
    style_header(ws, row=1, fills=fills)

    # 헤더 코멘트
    def hcomment(name, text):
        ws.cell(1, headers.index(name) + 1).comment = Comment(text, '시스템')
    hcomment('발문', 'HTML 허용')
    hcomment('해설', 'HTML 허용. <table> 가능.')
    hcomment('오답분석', '<ol><li>...</li></ol> 권장')
    hcomment('학습포인트', 'HTML 허용. <table> 가능.')

    # 샘플 3행
    base_meta = ['elec', None, None, 0, '독끝 전기기능사 필기', 'A',
                 '2026.05.29', '시스템', '2026-06-01',
                 '전기기능사', '2026년', '2회', None,
                 None, None, None, None, None,
                 None, None, '단답형', None,
                 None, None, None]
    samples = [
        (10001, 1, '전기이론', '전동기', '단상 유도전동기', '분상 기동형', '역회전 방법', SAMPLE_Q35),
        (10041, 41, '전기설비', '가공전선로', '경간 규제', '지지물별 경간', '철탑 경간', SAMPLE_Q41),
        (10002, 2, '전기이론', '정자계', '자성체', '자성체의 분류', '비투자율 분류', SAMPLE_Q02),
    ]
    for code, no, subj, chap, big, mid, content, data in samples:
        m = list(base_meta)
        m[1] = code              # 문항번호
        m[2] = f'elec{code}'     # 문항코드
        m[12] = no               # 번호
        m[13] = subj
        m[14] = chap
        m[15] = big
        m[16] = mid
        m[17] = content
        m[18] = 2                # 빈출도
        m[19] = 2                # 난이도
        m[22] = f'elec_A1_2026_02_{no:02d}_Q.PNG'
        m[23] = f'elec_A1_2026_02_{no:02d}_A.PNG'
        m[24] = f'elec_A1_2026_02_{no:02d}_E.PNG'
        row = m + [data[h] for h, _, _ in CONTENT_COLS]
        ws.append(row)

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 120
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')

    write_guide(wb)
    write_curriculum(wb)

    out = 'data/templates/[전기기능사 필기]CBT_문항관리_통합템플릿_v2.xlsx'
    wb.save(out)
    print(f'WRITE: {out}  ({ws.max_column} cols, {ws.max_row} rows)')


if __name__ == '__main__':
    build_simple()
    build_full()
    print('DONE')
