"""
v3 양식 — 가시성 강화 버전.

변경점 vs v2:
  - "번호" 컬럼: 문항 일련번호 (시험지 표시·정렬용)
  - 메타 영역 확장: 출처(5) + 코드(1) + 분류(5) + 속성(5) = 16컬럼
  - 해설·학습포인트는 단일 컬럼에 HTML 그대로 작성.
    · 그림은 본문 어디에나 `<img src="파일명">` 으로 삽입 (여러 개·교차 가능)
    · 오답분석도 HTML — 그림은 동일하게 <img src="…">

가시성 강화:
  - 1행: 영역 그룹 헤더 (출처 / 코드 / 분류 / 속성 / 발문 / 보기 / 정답 / 해설 / 오답분석 / 학습 POINT)
  - 2행: 컬럼명 (실제 헤더)
  - 영역별 진한 컬러 + 굵은 보더로 시각 구분
  - 컬럼 그룹화 (outline) — 메타 영역 접기/펼치기 가능
  - 데이터 유효성 검사 (DataValidation) — 드롭다운으로 입력 실수 방지
  - 행 높이 적정화 + 첫 두 행 고정 (Freeze panes)

출력:
  - data/templates/question-template-v3.xlsx
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

# ============== 스타일 ==============
WHITE = Font(color='FFFFFF', bold=True, size=12)
DARK = Font(color='1F2937', bold=True, size=11)
NORMAL = Font(color='1F2937', size=11)

THIN = Side(border_style='thin', color='CCCCCC')
THICK = Side(border_style='medium', color='4B5563')

INNER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GROUP_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

# 영역별 진한 컬러 (그룹 헤더용)
COLOR = {
    'meta':   '475569',   # slate-600
    'code':   '4F46E5',   # indigo-600
    'video':  '7C3AED',   # violet-600 (강의)
    'class':  '0891B2',   # cyan-600
    'attr':   'E11D48',   # rose-600
    'stem':   'DB2777',   # pink-600
    'opt':    '2563EB',   # blue-600
    'img':    'D97706',   # amber-600
    'ans':    '0D9488',   # teal-600
    'exp':    'EA580C',   # orange-600
    'wr':     '9333EA',   # purple-600
    'lp':     '059669',   # emerald-600
}

# 그룹 헤더 fill (진한 색)
def gfill(key):
    return PatternFill('solid', fgColor=COLOR[key])

# 컬럼 헤더 fill (연한 색)
LITE = {
    'meta':   'F1F5F9',
    'code':   'E0E7FF',
    'video':  'EDE9FE',
    'class':  'CFFAFE',
    'attr':   'FFE4E6',
    'stem':   'FCE7F3',
    'opt':    'DBEAFE',
    'img':    'FEF3C7',
    'ans':    'CCFBF1',
    'exp':    'FFEDD5',
    'wr':     'F3E8FF',
    'lp':     'D1FAE5',
}
def lfill(key):
    return PatternFill('solid', fgColor=LITE[key])


# ============== 컬럼 정의 ==============
# (헤더명, 너비, 카테고리키, 그룹라벨, 안내)
COLS = [
    # ----- 번호 (1) -----
    ('번호',          6, 'code',  '번호',     '문항 일련번호 (1, 2, 3 …). 시험지 표시·정렬용'),
    # ----- 출처 (5) -----
    ('과정',         11, 'meta',  '출처',     '자격증/과정명'),
    ('연도',          7, 'meta',  '출처',     '4자리 (2022)'),
    ('회차',          6, 'meta',  '출처',     '1·2·3·4'),
    ('사용교재',     20, 'meta',  '출처',     '기준 교재명'),
    ('교재구분',      8, 'meta',  '출처',     'A / B'),
    # ----- 코드 (1) -----
    ('문항코드',     22, 'code',  '코드',     '정렬·식별 키'),
    # ----- 강의 (1) -----
    ('강의주소',     28, 'video', '강의',     '시험 문제 풀이 시 상단에 표시될 강의 URL'),
    # ----- 분류 (5) -----
    ('과목ID',       11, 'class', '분류',     'theory / machinery / facility'),
    ('챕터',         14, 'class', '분류',     '대분류'),
    ('대유형',       18, 'class', '분류',     '중분류'),
    ('중유형',       18, 'class', '분류',     '소분류'),
    ('내용',         24, 'class', '분류',     '검색 키워드 / 좁은 분류'),
    # ----- 속성 (5) -----
    ('빈출도',        8, 'attr',  '속성',     '1~5 (★ 자주 나오는 정도)'),
    ('난이도',        8, 'attr',  '속성',     '1~5 (★ 어려움 정도)'),
    ('문제유형',     10, 'attr',  '속성',     '단답형 / 계산형 / 도식형 / 조합형 / 빈칸형 / 완성형'),
    ('변형이력',      8, 'attr',  '속성',     '0=원본, 1·2·… 변형'),
    ('비고',         18, 'attr',  '속성',     '운영자 메모'),
    # ----- 발문 -----
    ('발문',         45, 'stem',  '발문',     'HTML + LaTeX 수식 ($…$ / $$…$$)'),
    ('조건',         22, 'stem',  '발문',     '없으면 비움'),
    ('발문그림',     16, 'img',   '발문',     '파일명 / [필요: …]'),
    # ----- 보기 -----
    ('보기1',        24, 'opt',   '보기',     'HTML 허용'),
    ('보기1그림',    13, 'img',   '보기',     '파일명'),
    ('보기2',        24, 'opt',   '보기',     'HTML 허용'),
    ('보기2그림',    13, 'img',   '보기',     '파일명'),
    ('보기3',        24, 'opt',   '보기',     'HTML 허용'),
    ('보기3그림',    13, 'img',   '보기',     '파일명'),
    ('보기4',        24, 'opt',   '보기',     'HTML 허용'),
    ('보기4그림',    13, 'img',   '보기',     '파일명'),
    # ----- 정답 -----
    ('정답(1~4)',    8, 'ans',   '정답',     '1·2·3·4 중 하나'),
    # ----- 해설 (HTML + <img>) -----
    ('해설',         55, 'exp',   '해설',     'HTML + LaTeX($…$/$$…$$). 그림 <img src="파일명"> 본문 어디에나 (다중·교차 OK)'),
    # ----- 오답분석 (HTML + LaTeX + <img>) -----
    ('오답분석',     45, 'wr',    '오답분석', 'HTML <ol> 권장 + LaTeX 수식. 그림 <img src="파일명"> 삽입 가능'),
    # ----- 학습 POINT (HTML + LaTeX + <img>) -----
    ('학습포인트',   55, 'lp',    '학습POINT', 'HTML + LaTeX 수식. 그림 <img src="파일명"> 삽입 (공식·표와 교차 OK)'),
]


# ============== 블록 시퀀스 직렬화 (HTML-native) ==============
# 그림 블록은 <img src="파일명">. question_io.blocks_to_cell 와 동일 규칙.
def blocks_to_cell(blocks):
    parts = []
    for b in blocks:
        v = (b.get('value') or '').rstrip()
        if not v:
            continue
        if b['type'] == 'text':
            parts.append(v)
        elif b['type'] == 'image':
            parts.append(f'<img src="{v}">')
    return '\n\n'.join(parts)


# ============== 샘플 ==============
SAMPLES = [
    {  # 1번 콘덴서 — 단순
        '과정': '전기기능사', '연도': 2022, '회차': 1,
        '사용교재': '독끝 전기기능사 필기', '교재구분': 'A',
        '문항코드': 'elec_A_2022_01_01',
        '강의주소': 'https://youtu.be/sample-cap-q01',
        '과목ID': 'theory', '챕터': '정전기',
        '대유형': '콘덴서', '중유형': '정전용량', '내용': '정전용량 비례 관계',
        '빈출도': 4, '난이도': 2, '문제유형': '단답형', '변형이력': 0, '비고': '',
        '발문': '콘덴서의 정전용량에 대한 설명으로 틀린 것은?',
        '조건': '', '발문그림': '',
        '보기1': '전압에 반비례한다.', '보기1그림': '',
        '보기2': '이동 전하량에 비례한다.', '보기2그림': '',
        '보기3': '극판의 넓이에 비례한다.', '보기3그림': '',
        '보기4': '극판의 간격에 비례한다.', '보기4그림': '',
        '정답(1~4)': 4,
        '해설_블록': [
            {'type': 'text', 'value': '<p>정전용량 $C$는 극판의 넓이 $S$에 비례하고 극판의 간격 $d$에 반비례한다:</p>$$C = \\dfrac{\\varepsilon S}{d}\\ [\\mathrm{F}]$$<p>따라서 ④번 \'극판의 간격에 비례한다\'는 설명은 틀린 것이다.</p>'},
        ],
        '오답분석': '<ol><li>$C = \\dfrac{Q}{V}$ 관계에 의해 전압에 반비례한다.</li><li>$Q = CV$ 관계에 의해 전하량에 비례한다.</li><li>정전용량 공식 $C = \\dfrac{\\varepsilon S}{d}$ 에 의해 넓이에 비례한다.</li></ol>',
        '학습포인트_블록': [
            {'type': 'text', 'value': '<p><strong>정전용량 공식</strong></p>$$C = \\dfrac{\\varepsilon S}{d}\\ [\\mathrm{F}]$$<ul><li>$C$: 정전용량 $[\\mathrm{F}]$</li><li>$\\varepsilon$: 유전율 $[\\mathrm{F/m}]$</li><li>$S$: 극판 단면적 $[\\mathrm{m^2}]$</li><li>$d$: 극판 간격 $[\\mathrm{m}]$</li></ul>'},
        ],
    },
    {  # 21번 차동계전기 — 해설 3블록 시퀀스
        '과정': '전기기능사', '연도': 2022, '회차': 1,
        '사용교재': '독끝 전기기능사 필기', '교재구분': 'A',
        '문항코드': 'elec_A_2022_01_21',
        '강의주소': 'https://youtu.be/sample-differential-relay',
        '과목ID': 'machinery', '챕터': '변압기',
        '대유형': '보호 계전기', '중유형': '계전기 종류', '내용': '내부 고장 보호용 계전기',
        '빈출도': 5, '난이도': 2, '문제유형': '단답형', '변형이력': 0, '비고': '',
        '발문': '변압기 내부 고장 보호에 쓰이는 계전기로써 가장 알맞은 것은?',
        '조건': '', '발문그림': '',
        '보기1': '차동계전기', '보기1그림': '',
        '보기2': '접지계전기', '보기2그림': '',
        '보기3': '과전류계전기', '보기3그림': '',
        '보기4': '역상계전기', '보기4그림': '',
        '정답(1~4)': 1,
        '해설_블록': [
            {'type': 'text', 'value': '<p>변압기 내부 권선의 층간 단락이나 지락 사고 시, 입력 전류와 출력 전류 사이에 차이가 발생한다.</p>'},
            {'type': 'image', 'value': 'q21_differential_diagram.png'},
            {'type': 'text', 'value': '<p>이 차이(차전류)를 양쪽 변류기(CT) 회로에서 검출한다.</p>'},
            {'type': 'image', 'value': 'q21_ct_balance.png'},
            {'type': 'text', 'value': '<p>차전류가 정정값을 넘으면 동작하는 <strong>차동 계전기</strong>가 가장 적합하다.</p>'},
        ],
        '오답분석': '<ol><li value="2"><strong>접지계전기</strong>: 영상 전류를 검출하여 지락 사고 시 동작.</li><li><strong>과전류계전기</strong>: 설정값 이상 전류 시 동작.</li><li><strong>역상계전기</strong>: 역상 전류 검출.</li></ol><p>각 계전기의 보호 범위 비교:</p><img src="q21_relay_compare.png">',
        '학습포인트_블록': [
            {'type': 'text', 'value': '<p><strong>주요 보호 계전기</strong></p><table><thead><tr><th>계전기</th><th>보호 대상</th></tr></thead><tbody><tr><td>차동계전기</td><td>변압기·발전기 내부 권선 고장</td></tr><tr><td>접지(지락)계전기</td><td>지락 사고 (영상 전류)</td></tr><tr><td>과전류계전기</td><td>과부하·단락 (설정값 초과 전류)</td></tr><tr><td>역상계전기</td><td>불평형·결상 (역상 전류)</td></tr></tbody></table>'},
        ],
    },
    {  # 51번 그림기호 — 발문그림 [필요] 마커 + 학습포인트 3블록
        '과정': '전기기능사', '연도': 2022, '회차': 1,
        '사용교재': '독끝 전기기능사 필기', '교재구분': 'A',
        '문항코드': 'elec_A_2022_01_51',
        '강의주소': '',
        '과목ID': 'facility', '챕터': '전기 도면',
        '대유형': '접점 기호', '중유형': '수동 조작 접점', '내용': '그림기호 식별',
        '빈출도': 3, '난이도': 2, '문제유형': '도식형', '변형이력': 0,
        '비고': '발문그림 필요 — 수동 업로드',
        '발문': '아래 그림기호가 나타내는 것은?',
        '조건': '',
        '발문그림': '[필요: 기호심볼 — 수동조작 접점]',
        '보기1': '한시 계전기 접점', '보기1그림': '',
        '보기2': '전자 접속기 접점', '보기2그림': '',
        '보기3': '수동 조작 접점', '보기3그림': '',
        '보기4': '조작 개폐기 잔류 접점', '보기4그림': '',
        '정답(1~4)': 3,
        '해설_블록': [
            {'type': 'text', 'value': '그림의 기호는 사용자가 손으로 직접 조작하여 회로를 개폐하는 <strong>수동 조작 접점</strong>을 나타낸다.'},
        ],
        '오답분석': '',
        '학습포인트_블록': [
            {'type': 'text', 'value': '<p><strong>접점 기호 분류</strong></p>'},
            {'type': 'image', 'value': '[필요: 접점 기호 4종 비교도]'},
            {'type': 'text', 'value': '<table><thead><tr><th>접점</th><th>특징</th></tr></thead><tbody><tr><td>한시 접점</td><td>설정 시간 경과 후 동작 (Timer)</td></tr><tr><td>전자 접속기 접점</td><td>전자코일 여자에 의해 동작 (MC)</td></tr><tr><td>수동 조작 접점</td><td>손으로 직접 개폐</td></tr><tr><td>잔류 접점</td><td>조작 후에도 접점 상태 유지</td></tr></tbody></table>'},
        ],
    },
]


# ============== 사용가이드 ==============
GUIDE_ROWS = [
    ('컬럼', '설명', '필수', '예시'),
    ('번호', '문항 일련번호 (1, 2, 3 …). 시험지 표시·정렬용', '권장', '1'),
    ('과정', '자격증 / 과정명', '필수', '전기기능사'),
    ('연도', '시험 연도 (4자리)', '필수', '2022'),
    ('회차', '시험 회차 (1·2·3·4)', '필수', '1'),
    ('사용교재', '기준 교재명', '권장', '독끝 전기기능사 필기'),
    ('교재구분', '교재 판/구분 코드', '권장', 'A'),
    ('문항코드', '정렬·식별 키. 규칙: elec_<교재>_<연도>_<회차>_<번호>', '필수', 'elec_A_2022_01_01'),
    ('강의주소', '시험 문제 풀이 시 상단 강의 버튼이 연결할 URL', '선택', 'https://youtu.be/...'),
    ('과목ID', 'theory / machinery / facility', '필수', 'theory'),
    ('챕터', '대분류', '권장', '정자계'),
    ('대유형', '중분류', '권장', '자성체와 자기회로'),
    ('중유형', '소분류', '권장', '투자율과 자성체 분류'),
    ('내용', '가장 좁은 분류 (검색 키워드)', '권장', '자성체의 분류 (강·상·반자성)'),
    ('빈출도', '1~5 정수 (자주 나오는 정도)', '권장', '4'),
    ('난이도', '1~5 정수 (어려움 정도)', '권장', '2'),
    ('문제유형', '단답형 / 계산형 / 도식형 / 조합형 / 빈칸형 / 완성형', '권장', '단답형'),
    ('변형이력', '0=원본, 1·2·… 변형 차수', '선택', '0'),
    ('비고', '운영자 메모 (자유)', '선택', '변형문제, 그림 보강 필요'),
    ('발문', '문제 본문 (HTML + LaTeX 수식)', '필수', '$\\mu_s$ 는 ...'),
    ('조건', '주어진 조건 (없으면 비움)', '선택', 'P=100[Ω]'),
    ('발문그림', '발문에 들어가는 그림 파일명 / [필요: …]', '선택', 'q01_stem.png'),
    ('보기1~4', '4지선다 본문 (HTML + LaTeX 수식)', '필수', '$V_l = \\sqrt{3}\\,V_p$'),
    ('보기1~4그림', '보기가 그림형일 때 파일명', '선택', 'opt_q35_1.png'),
    ('정답(1~4)', '1·2·3·4 중 하나', '필수', '4'),
    ('해설', 'HTML + LaTeX 수식. 그림은 <img src="파일명">을 본문 어디에나 삽입 (여러 개·텍스트와 교차 가능)', '필수', '<p>$C=\\dfrac{Q}{V}$ 이므로…</p><img src="q01_a.png"><p>이어서</p>'),
    ('오답분석', 'HTML + LaTeX 수식. 그림 필요 시 <img src="파일명"> 삽입 가능 (<ol><li> 권장)', '선택', '<ol><li>$Q=CV$ 이므로 …</li></ol>'),
    ('학습포인트', 'HTML + LaTeX 수식. 그림은 <img src="파일명"> 삽입 (공식·표와 교차 가능)', '선택', '<p>공식</p>$$C=\\dfrac{\\varepsilon S}{d}$$'),
    ('', '', '', ''),
    ('[해설 / 오답분석 / 학습포인트 작성 규칙 — HTML-native]', '', '', ''),
    ('규칙 1', '셀에 HTML을 그대로 작성한다. 여러 줄·표·리스트 모두 가능.', '', ''),
    ('규칙 2', '그림은 본문 안에 <img src="파일명"> 으로 넣는다. 위치 제한 없음.', '', '<img src="q21_diagram.png">'),
    ('규칙 3', '그림을 여러 개 넣거나 텍스트와 번갈아 배치해도 된다 (해설↔그림↔해설↔그림 …). 등록·시험 화면이 자동으로 블록 분리·렌더한다.', '', ''),
    ('규칙 4', '그림 파일이 아직 없으면 <img src="[필요: 설명]"> 형태로 마커.', '', '<img src="[필요: TRIAC 심볼]">'),
    ('규칙 5', '구 표기 [IMG: 파일명] 도 계속 인식한다 (기존 파일 호환).', '', '[IMG: q21_diagram.png]'),
    ('', '', '', ''),
    ('[HTML 허용 태그]', '', '', ''),
    ('인라인', '<i> <em> <strong> <u> <sub> <sup> <br> <code>', '', ''),
    ('블록', '<p> <ul> <ol> <li> <table> <thead> <tbody> <tr> <th> <td>', '', ''),
    ('이미지', '<img src="파일명">  — 해설/오답분석/학습포인트 본문 어디에나', '', ''),
    ('특수문자', '&lt;는 <, &gt;는 >, &amp;는 &', '', ''),
    ('', '', '', ''),
    ('[수식 — LaTeX (KaTeX)]', '', '', ''),
    ('인라인 수식', '$ ... $ 로 감싼다. 글 안에 섞어 쓰기.', '', '저항 $R = \\dfrac{V}{I}$ 이다'),
    ('디스플레이 수식', '$$ ... $$ 로 감싼다. 가운데 정렬·큰 글씨(독립 줄).', '', '$$P = VI = I^2 R$$'),
    ('분수·루트·첨자', '\\dfrac{a}{b} · \\sqrt{x} · x^{2} · x_{1} · \\varepsilon \\pi \\theta', '', '$\\dfrac{1}{2\\pi f C}$ · $\\sqrt{3}$'),
    ('단위', '\\mathrm{} 으로 정체 표기 권장.', '', '$[\\mathrm{F}]$ · $[\\mathrm{m^2}]$'),
    ('주의', '본문 글자로 $ 를 쓰려면 \\$ 로 이스케이프. < > 흉내(<i><sub>)보다 수식은 LaTeX 권장.', '', ''),
]


def set_print_landscape(ws, paper_size=8):
    """가로 A3 (paperSize=8) / 가로 1페이지에 맞춤."""
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.paperSize = paper_size
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.3, bottom=0.3)
    ws.print_options.horizontalCentered = True


def build_workbook(out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '문항등록'

    n_cols = len(COLS)

    # ----- 1행: 그룹 헤더 (병합) -----
    # 같은 group_label이 연속되는 구간을 병합
    headers = [c[0] for c in COLS]
    widths = [c[1] for c in COLS]
    tones = [c[2] for c in COLS]
    groups = [c[3] for c in COLS]
    hints = [c[4] for c in COLS]

    i = 0
    while i < n_cols:
        g = groups[i]
        j = i
        while j + 1 < n_cols and groups[j + 1] == g:
            j += 1
        # i..j 병합, 라벨 = g
        ws.cell(1, i + 1, g)
        if j > i:
            ws.merge_cells(start_row=1, start_column=i + 1, end_row=1, end_column=j + 1)
        # 그룹 헤더 스타일 — 첫 컬럼의 톤 색상 사용
        cell = ws.cell(1, i + 1)
        cell.fill = gfill(tones[i])
        cell.font = WHITE
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 굵은 보더는 그룹 양 끝
        for k in range(i, j + 1):
            ws.cell(1, k + 1).border = Border(
                top=THICK,
                bottom=THICK,
                left=THICK if k == i else THIN,
                right=THICK if k == j else THIN,
            )
        i = j + 1

    ws.row_dimensions[1].height = 26

    # ----- 2행: 컬럼 헤더 -----
    for c, (name, width, tone, group, hint) in enumerate(COLS, 1):
        cell = ws.cell(2, c, name)
        cell.fill = lfill(tone)
        cell.font = DARK
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = INNER
        cell.comment = Comment(hint, '시스템')
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.row_dimensions[2].height = 32

    # ----- 컬럼 그룹화 (outline) — 출처/분류/속성은 접을 수 있게 -----
    # 그룹 라벨 기준으로 계산 (컬럼 추가/이동에 안전)
    _collapsible = {'출처', '분류', '속성'}
    for c, (_n, _w, _t, group, _h) in enumerate(COLS, 1):
        if group in _collapsible:
            ws.column_dimensions[get_column_letter(c)].outlineLevel = 1

    # ----- 3행~ 샘플 데이터 -----
    def col_idx(name):
        return headers.index(name) + 1

    for s in SAMPLES:
        row = []
        for header, _, _, _, _ in COLS:
            if header == '해설':
                row.append(blocks_to_cell(s.get('해설_블록', [])))
            elif header == '학습포인트':
                row.append(blocks_to_cell(s.get('학습포인트_블록', [])))
            else:
                row.append(s.get(header, ''))
        ws.append(row)

    # 본문 셀 wrap + 행 높이
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 160
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # 컬럼 영역색 약하게
            cell.border = INNER

    # ----- 데이터 유효성 (드롭다운) -----
    last_row = max(ws.max_row + 200, 500)   # 향후 입력 행도 적용

    def add_dv(col_name, formula, prompt=None):
        c = col_idx(col_name)
        col = get_column_letter(c)
        dv = DataValidation(type='list', formula1=formula, allow_blank=True)
        if prompt:
            dv.promptTitle = col_name
            dv.prompt = prompt
        dv.showDropDown = False
        ws.add_data_validation(dv)
        dv.add(f'{col}3:{col}{last_row}')

    add_dv('과정', '"전기기능사,전기산업기사,전기기사"')
    add_dv('과목ID', '"theory,machinery,facility"')
    add_dv('문제유형', '"단답형,계산형,도식형,조합형,빈칸형,완성형"')
    add_dv('정답(1~4)', '"1,2,3,4"')
    add_dv('빈출도', '"1,2,3,4,5"')
    add_dv('난이도', '"1,2,3,4,5"')

    # ----- Freeze panes: 첫 두 행 + 문항코드 컬럼까지 고정 -----
    code_col = col_idx('문항코드')
    ws.freeze_panes = ws.cell(3, code_col + 1)

    # 가로 A3 한 페이지에 맞춤 (PDF 변환·인쇄 시)
    set_print_landscape(ws, paper_size=8)

    # ===== 사용가이드 시트 =====
    g = wb.create_sheet('사용가이드')
    for row in GUIDE_ROWS:
        g.append(row)
    for c in range(1, 5):
        cell = g.cell(1, c)
        cell.font = WHITE
        cell.fill = PatternFill('solid', fgColor=COLOR['meta'])
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = INNER
    g.row_dimensions[1].height = 28
    for col_letter, w in zip('ABCD', [22, 75, 8, 50]):
        g.column_dimensions[col_letter].width = w
    for r in range(2, g.max_row + 1):
        for c in range(1, 5):
            g.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    set_print_landscape(g, paper_size=9)  # A4 가로

    # ===== 커리큘럼ID 시트 =====
    c = wb.create_sheet('커리큘럼ID')
    rows = [
        ('과목ID', '과목명', '챕터(예시)'),
        ('theory', '전기이론', '직류회로 / 정전기 / 정자계 / 교류회로 / 전기 기초'),
        ('machinery', '전기기기', '직류기 / 동기기 / 변압기 / 유도기 / 정류기'),
        ('facility', '전기설비', '옥내배선 / 가공전선로 / 접지 / 보호장치 / 시공'),
    ]
    for r in rows:
        c.append(r)
    for cc in range(1, 4):
        cell = c.cell(1, cc)
        cell.font = WHITE
        cell.fill = PatternFill('solid', fgColor=COLOR['meta'])
        cell.border = INNER
        cell.alignment = Alignment(horizontal='center')
    for col_letter, w in zip('ABC', [12, 14, 80]):
        c.column_dimensions[col_letter].width = w

    # ===== 색상 범례 시트 =====
    leg = wb.create_sheet('색상범례')
    legend_rows = [
        ('영역', '컬러', '컬럼', '용도'),
        ('번호', '인디고', '번호', '문항 일련번호 (시험지 표시·정렬용)'),
        ('출처', '회색', '과정·연도·회차·사용교재·교재구분', '시험 출처 정보'),
        ('코드', '인디고', '문항코드', '정렬·식별 키'),
        ('강의', '바이올렛', '강의주소', '문제 풀이 화면 상단 강의 버튼 URL'),
        ('분류', '시안', '과목ID·챕터·대유형·중유형·내용', '카테고리 (세로축)'),
        ('속성', '로즈', '빈출도·난이도·문제유형·변형이력·비고', '문항 메타 속성 (가로축)'),
        ('발문', '핑크', '발문·조건', '문제 본문'),
        ('그림', '앰버', '*그림 컬럼들', '파일명 또는 [필요: …] 마커'),
        ('보기', '블루', '보기1~4', '4지선다'),
        ('정답', '틸', '정답(1~4)', '1·2·3·4 중 하나'),
        ('해설', '오렌지', '해설', 'HTML + LaTeX($…$) + <img> (다중·교차 OK)'),
        ('오답분석', '퍼플', '오답분석', 'HTML <ol> + LaTeX + <img>'),
        ('학습 POINT', '에메랄드', '학습포인트', 'HTML + LaTeX + <img> (공식·표·그림 교차)'),
    ]
    for r in legend_rows:
        leg.append(r)
    tones_for_legend = [None, 'code', 'meta', 'code', 'video', 'class', 'attr', 'stem', 'img', 'opt', 'ans', 'exp', 'wr', 'lp']
    for r, t in enumerate(tones_for_legend, 1):
        for c in range(1, 5):
            cell = leg.cell(r, c)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = INNER
            if r == 1:
                cell.fill = PatternFill('solid', fgColor=COLOR['meta'])
                cell.font = WHITE
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif t:
                if c == 1:
                    cell.fill = gfill(t)
                    cell.font = WHITE
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c == 2:
                    cell.fill = lfill(t)
    for col_letter, w in zip('ABCD', [16, 12, 42, 36]):
        leg.column_dimensions[col_letter].width = w
    leg.row_dimensions[1].height = 24
    set_print_landscape(leg, paper_size=9)  # A4 가로

    wb.save(out_path)
    print(f'WROTE: {out_path}')
    print(f'  컬럼 {ws.max_column}개, 샘플 행 {ws.max_row - 2}개')
    print(f'  시트: {wb.sheetnames}')


if __name__ == '__main__':
    build_workbook('data/templates/question-template-v3.xlsx')
