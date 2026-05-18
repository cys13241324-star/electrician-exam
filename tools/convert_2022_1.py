"""2022년 1회 PDF → v2 엑셀 변환.
- 자동 파싱 가능한 부분만 채우고, 누락/의심 부분은 '이슈' 컬럼과 별도 시트에 기록.
- 출력: data/2022_1회_변환결과.xlsx
"""
import sys, re, os
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pypdf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

try:
    from manual_fixes import FIXES, IMAGE_NEEDS
except ImportError:
    FIXES, IMAGE_NEEDS = {}, {}

PDF = r'data/templates/sample/2022년 1회.pdf'
OUT = r'data/2022_1회_변환결과.xlsx'

# 정답 인코딩 변환
ANS_MAP = {'➊': 1, '➋': 2, '➌': 3, '➍': 4,
           '①': 1, '②': 2, '③': 3, '④': 4}

# 카테고리 매핑 (페이지 헤더에서 추출되는 명칭 → 과목ID)
SUBJECT_MAP = {
    '전기이론': 'theory',
    '전기 기기': 'machinery',
    '전기기기': 'machinery',
    '전기 설비': 'facility',
    '전기설비': 'facility',
}

# ---------- 1. PDF 전체 텍스트 (페이지 2부터, page 1 스킵) ----------
def load_pages():
    r = pypdf.PdfReader(PDF)
    pages = []
    for i, p in enumerate(r.pages, 1):
        try:
            t = p.extract_text() or ''
        except Exception as e:
            t = ''
        pages.append((i, t))
    return pages

# ---------- 2. 페이지에서 노이즈 라인 제거 ----------
NOISE_PATTERNS = [
    re.compile(r'^\s*\d+\s*전기기능사 필기 기본서.*$'),       # 좌측페이지 헤더
    re.compile(r'^\s*PART\s*\d+\s*8개년 기출문제.*$'),          # 우측페이지 헤더
    re.compile(r'^\s*2022년.*회 CBT 모의고사\s*$'),
    re.compile(r'^\s*\d+\s*문항\s*$'),                          # "10문항", "19문항"
    re.compile(r'^\s*1회독\(\d+문항\).*$'),
    re.compile(r'^\s*\*\s*회차별 정답수.*$'),
    re.compile(r'^\s*이론\s*$'),
    re.compile(r'^\s*기기\s*$'),
    re.compile(r'^\s*설비\s*$'),
    re.compile(r'^\s*합계\s*$'),
]

CATEGORY_LINE = re.compile(r'^\s*(전기\s*이론|전기\s*기기|전기\s*설비)\s*$')
QNUM_LINE = re.compile(r'^\s*(\d{2})\s*$')                    # "01", "02", ...
END_TAG_LINE = re.compile(r'^\[(전기이론|전기기기|전기설비)\]\s*\(\d+p\)\s*([➊➋➌➍])\s*$')

def normalize_subject(s):
    s = s.replace(' ', '')
    return SUBJECT_MAP.get(s, None) or SUBJECT_MAP.get(s.replace('', ''), None)

# ---------- 3. 문항 단위 파싱 ----------
def parse_questions(pages):
    """Return list of dict per question."""
    # 페이지 1은 이전 회차 60번이라 스킵
    all_lines = []
    for pno, text in pages:
        if pno == 1:
            continue
        for ln in text.splitlines():
            all_lines.append((pno, ln))

    # 노이즈 제거
    clean = []
    for pno, ln in all_lines:
        s = ln.strip()
        if not s:
            clean.append((pno, ''))
            continue
        skip = any(p.match(s) for p in NOISE_PATTERNS)
        if skip:
            continue
        clean.append((pno, ln))

    # 카테고리 트래킹 + 문항 분리
    questions = []
    current_subject = None
    current_q = None

    i = 0
    while i < len(clean):
        pno, ln = clean[i]
        s = ln.strip()

        # 카테고리 헤더
        m = CATEGORY_LINE.match(s)
        if m:
            current_subject = normalize_subject(m.group(1))
            i += 1
            continue

        # 문항 시작 (2자리 숫자 단독 라인)
        m = QNUM_LINE.match(s)
        if m:
            no = int(m.group(1))
            # 이전 문항 종료
            if current_q is not None:
                questions.append(current_q)
            current_q = {
                'no': no,
                'subject': current_subject,
                'page_start': pno,
                'raw_lines': [],
            }
            i += 1
            continue

        # 일반 라인은 현재 문항에 추가
        if current_q is not None:
            current_q['raw_lines'].append((pno, ln))
        i += 1

    if current_q is not None:
        questions.append(current_q)

    return questions

# ---------- 4. 각 문항의 raw_lines를 구조화 ----------
OPT_INLINE = re.compile(r'(?<![가-힣A-Za-z0-9])([①②③④])')  # ① 마커 식별

def split_options(text):
    """텍스트 내 ①②③④ 마커 기준 분리. 마커 순서대로 4개 보기 반환.
       텍스트 자체가 비어있거나 마커가 4개 미만이면 부분 결과."""
    parts = {}
    matches = list(OPT_INLINE.finditer(text))
    # 정답 라벨용 ①②③④ 4개만 추적 (보기 안 ①②는 별개 처리 필요하나 일단 1차 추출)
    used = []
    for m in matches:
        ch = m.group(1)
        if ch in used:
            continue
        if len(used) >= 4:
            break
        used.append(ch)
        start = m.end()
        end = matches[matches.index(m) + 1].start() if matches.index(m) + 1 < len(matches) else len(text)
        # 다음 보기 마커까지
        next_marker_pos = None
        for n in matches[matches.index(m) + 1:]:
            if n.group(1) not in used:
                next_marker_pos = n.start()
                break
        if next_marker_pos is None:
            next_marker_pos = len(text)
        parts[ch] = text[start:next_marker_pos].strip()
    return [parts.get(c, '') for c in '①②③④']

PUA_RE = re.compile(r'[-]+')

def replace_pua(s):
    """PDF 폰트 매핑 실패로 PUA에 들어간 글자를 [수식] 마커로 치환."""
    if not s:
        return s
    return PUA_RE.sub('[수식]', s)

# 자모 분리 패턴: 한글 단글자 + 공백 + 한글 단글자가 3회 이상 연속할 때만 매치
JAMO_SPLIT_RE = re.compile(r'((?:[가-힣]\s){2,}[가-힣])')

def collapse_jamo_strict(s):
    """`기 본 파` 같은 자모 분리 패턴만 골라서 붙임. 일반 어절 공백은 유지."""
    if not s:
        return s, False
    changed = False
    def repl(m):
        nonlocal changed
        changed = True
        return re.sub(r'\s+', '', m.group(1))
    new = JAMO_SPLIT_RE.sub(repl, s)
    return new, changed

# 학습 POINT 시작 라인 패턴 (해설 본문 다음에 오는 공식/법칙/특징 박스 헤더)
LP_HEADER_RE = re.compile(
    r'^.{2,30}(공식|법칙|특징|분류|종류|성질|관계|표현식|효율|구분|반작용|기준|구조|길이|깊이|주파수|표준|화학반응|순시값)\s*$'
)

def split_explain_lp(explain):
    """해설 본문을 (해설, 학습포인트)로 분리.
       박스 헤더 라인을 찾아 그 이전은 해설, 이후는 학습포인트."""
    if not explain:
        return '', ''
    lines = explain.split('\n')
    lp_start = None
    for i, ln in enumerate(lines):
        s = ln.strip()
        if not s:
            continue
        if LP_HEADER_RE.match(s) and i >= 2:
            lp_start = i
            break
    if lp_start is None:
        return explain, ''
    body = '\n'.join(lines[:lp_start]).strip()
    lp = '\n'.join(lines[lp_start:]).strip()
    return body, lp

# 발문/해설에서 그림이 필요한 신호
def detect_image_need(text):
    if not text:
        return ''
    if re.search(r'그림과\s*같은', text):
        return '[필요: 회로도/도면]'
    if re.search(r'그림기호', text):
        return '[필요: 기호 심볼]'
    if re.search(r'아래\s*그림', text):
        return '[필요: 도면/그림]'
    if re.search(r'다음\s*그림', text):
        return '[필요: 도면/그림]'
    if re.search(r'심볼', text):
        return '[필요: 심볼]'
    return ''

# 빈칸 표시 `( ① )` → `(빈칸1)` 로 사전 치환 (보기 마커와 충돌 방지)
BLANK_FILL_RE = re.compile(r'\(\s*([①②③④])\s*\)')
BLANK_MAP = {'①': '(빈칸1)', '②': '(빈칸2)', '③': '(빈칸3)', '④': '(빈칸4)'}

def preprocess_blanks(lines):
    """발문에 ( ① ) 같은 빈칸이 있으면 (빈칸1)로 치환하고 빈칸문제 플래그 반환."""
    has_blank = False
    out = []
    for ln in lines:
        if BLANK_FILL_RE.search(ln):
            has_blank = True
            ln = BLANK_FILL_RE.sub(lambda m: BLANK_MAP.get(m.group(1), m.group(0)), ln)
        out.append(ln)
    return out, has_blank

def structure_question(q):
    """raw_lines → 발문/보기/해설/오답분석/정답/이슈 추출."""
    raw_lines = [ln for _, ln in q['raw_lines']]
    lines, has_blank = preprocess_blanks(raw_lines)
    full = '\n'.join(lines)

    issues = []
    if has_blank:
        issues.append('빈칸문제 (보기에 ①②가 다시 등장)')

    # 정답 추출: 끝 태그 [전기이론] (00p) ➊
    answer = None
    tag_idx = -1
    for idx, ln in enumerate(lines):
        m = END_TAG_LINE.match(ln.strip())
        if m:
            answer = ANS_MAP.get(m.group(2))
            tag_idx = idx
            # 카테고리 태그도 잡되, PDF 오류로 전부 '전기이론'이므로 무시
            break
    if answer is None:
        issues.append('정답 추출 실패')

    body = lines[:tag_idx] if tag_idx >= 0 else lines

    # 발문 / 보기 / 해설 / 오답분석 분리
    # - 발문 = 첫 ① 이전까지
    # - 보기 = 첫 ① ~ 보기 ④ 끝까지 (※ 보기 본문에 ①②가 다시 나오는 빈칸문제 주의)
    # - 해설 = 보기 ④ 끝 ~ '오답분석' 라인 이전
    # - 오답분석 = '오답분석' 라인 이후
    first_opt_line = None
    for idx, ln in enumerate(body):
        if OPT_INLINE.search(ln):
            first_opt_line = idx
            break

    stem_lines = body[:first_opt_line] if first_opt_line is not None else body
    rest = body[first_opt_line:] if first_opt_line is not None else []

    # 오답분석 라인 찾기
    odab_idx = None
    for idx, ln in enumerate(rest):
        if ln.strip().startswith('오답분석'):
            odab_idx = idx
            break

    if odab_idx is not None:
        opt_and_explain = rest[:odab_idx]
        odab_lines = rest[odab_idx + 1:]
    else:
        opt_and_explain = rest
        odab_lines = []

    # 보기 ~ 해설 분리
    # 휴리스틱: 첫 ① 시작 후 ④까지의 텍스트를 보기로, 그 다음 빈 줄 또는 ' ' 단독 라인 이후를 해설로
    opt_text_block = []
    explain_block = []
    found_4 = False
    in_options = True
    for ln in opt_and_explain:
        if in_options:
            opt_text_block.append(ln)
            if '④' in ln:
                found_4 = True
            elif found_4 and ln.strip() == '':
                in_options = False
            elif found_4 and not OPT_INLINE.search(ln) and ln.strip() and len(opt_text_block) > 1:
                # ④ 이후 빈 줄 없이 바로 해설 시작
                # 짧은 토큰만 있는 라인이면 보기 잔여로 간주 (수식)
                if len(ln.strip()) > 10:
                    in_options = False
                    explain_block.append(opt_text_block.pop())  # 마지막 라인을 해설로 이동
        else:
            explain_block.append(ln)

    opt_text = '\n'.join(opt_text_block)
    options = split_options(opt_text)

    # 빈 보기 이슈
    for i, o in enumerate(options, 1):
        if not o or len(o.strip()) < 1:
            issues.append(f'보기{i} 비어 있음(수식/그림)')

    # 보기 자모 공백 정리 (1회만 이슈 표시)
    options_clean = []
    any_jamo = False
    for o in options:
        c, ch = collapse_jamo_strict(o)
        c = replace_pua(c)
        if ch:
            any_jamo = True
        options_clean.append(c)
    options = options_clean
    # 자모 분리 정리는 정상 동작 — 이슈가 아니라 정보 플래그
    jamo_cleaned = any_jamo

    # 발문 / 해설 / 오답분석 처리
    def post(lines):
        txt = '\n'.join(lines).strip()
        txt, _ = collapse_jamo_strict(txt)
        txt = replace_pua(txt)
        # 단위 공백 정리
        txt = re.sub(r'\bkV\s+A\b', 'kVA', txt)
        return txt

    stem = post(stem_lines)
    explain = post(explain_block)
    odab = post(odab_lines)

    # 해설을 (해설, 학습포인트)로 자동 분리 — 정상 동작은 이슈가 아니라 정보
    explain, learning_point = split_explain_lp(explain)
    lp_auto = bool(learning_point)

    # 그림 필요 위치 자동 감지
    stem_img = detect_image_need(stem)
    explain_img = detect_image_need(explain) or detect_image_need(learning_point)

    # 빈 변수 괄호 패턴 ()
    if re.search(r'\(\s*\)', stem + explain + odab):
        issues.append('수식 변수 누락 의심 (빈 괄호 ())')

    # 그림 키워드
    if re.search(r'그림|심볼', stem):
        issues.append('발문 그림 포함 (수동 처리)')

    # 표 키워드 (해설/오답분석에 표 추정 패턴: 셀구분되는 짧은 항목 3개 이상)
    if re.search(r'(공식|법칙|표준|특징|길이|깊이|주파수|반작용)\s*\n', explain) and explain.count('\n') > 8:
        issues.append('해설에 표/공식블록 포함 가능 (수동 정리 권장)')

    # 텍스트가 너무 짧으면 추출 실패 의심
    if len(stem) < 8:
        issues.append('발문이 너무 짧음 (추출 실패 의심)')
    if len(explain) < 8:
        issues.append('해설이 너무 짧음 (추출 실패 의심)')

    return {
        'no': q['no'],
        'subject': q['subject'],
        'page_start': q['page_start'],
        'stem': stem,
        'stem_img': stem_img,
        'opt1': options[0], 'opt1_img': '',
        'opt2': options[1], 'opt2_img': '',
        'opt3': options[2], 'opt3_img': '',
        'opt4': options[3], 'opt4_img': '',
        'answer': answer,
        'explain': explain,
        'explain_img': explain_img,
        'odab': odab,
        'learning_point': learning_point,
        'lp_img': '',
        'issues': '; '.join(issues),
        'lp_auto': lp_auto,
        'jamo_cleaned': jamo_cleaned,
        'manual_applied': [],
    }

# ---------- manual_fixes 적용 ----------
FIELD_TO_KEY = {
    '발문': 'stem',
    '발문그림': 'stem_img',
    '보기1': 'opt1', '보기1그림': 'opt1_img',
    '보기2': 'opt2', '보기2그림': 'opt2_img',
    '보기3': 'opt3', '보기3그림': 'opt3_img',
    '보기4': 'opt4', '보기4그림': 'opt4_img',
    '정답(1~4)': 'answer',
    '해설': 'explain',
    '해설그림': 'explain_img',
    '오답분석': 'odab',
    '학습포인트': 'learning_point',
    '학습포인트그림': 'lp_img',
}

def apply_manual_fixes(q):
    """FIXES / IMAGE_NEEDS 의 값으로 컬럼 덮어쓰기. 적용 컬럼은 manual_applied에 기록."""
    applied = []
    for src in (FIXES.get(q['no'], {}), IMAGE_NEEDS.get(q['no'], {})):
        for col_name, val in src.items():
            key = FIELD_TO_KEY.get(col_name)
            if key is None or val is None or val == '':
                continue
            q[key] = val
            applied.append(col_name)
    q['manual_applied'] = applied
    # 수동 보완으로 해결된 진짜 이슈는 제거
    if applied:
        cleaned = []
        for it in (q.get('issues') or '').split('; '):
            it = it.strip()
            if not it:
                continue
            # 보기 비어있음 / 발문 그림 포함 / 정답 추출 실패는 수동보완으로 해결됨
            if (re.match(r'보기\d+ 비어', it) and any('보기' in a for a in applied)):
                continue
            if it.startswith('발문 그림 포함') and '발문그림' in applied:
                continue
            if it == '정답 추출 실패' and '정답(1~4)' in applied:
                continue
            cleaned.append(it)
        q['issues'] = '; '.join(cleaned)
    return q

# ---------- 5. 엑셀로 출력 (v2 단순 템플릿 + 이슈 컬럼) ----------
HDR_FILL = PatternFill('solid', fgColor='FFE4EC')
META_FILL = PatternFill('solid', fgColor='E8F0FE')
CONTENT_FILL = PatternFill('solid', fgColor='FFF7E0')
IMG_FILL = PatternFill('solid', fgColor='E8F8E8')
ANS_FILL = PatternFill('solid', fgColor='FFE0E0')
ISSUE_FILL = PatternFill('solid', fgColor='FFD6D6')
HDR_FONT = Font(bold=True, size=11)
THIN = Side(border_style='thin', color='CCCCCC')
HDR_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# v2 단순 컬럼 + 이슈 + 페이지
COLUMNS = [
    ('번호', 6, META_FILL),
    ('과목ID', 12, META_FILL),
    ('토픽ID', 14, META_FILL),
    ('서브토픽ID', 14, META_FILL),
    ('발문', 50, CONTENT_FILL),
    ('조건', 18, CONTENT_FILL),
    ('발문그림', 14, IMG_FILL),
    ('보기1', 28, CONTENT_FILL),
    ('보기1그림', 12, IMG_FILL),
    ('보기2', 28, CONTENT_FILL),
    ('보기2그림', 12, IMG_FILL),
    ('보기3', 28, CONTENT_FILL),
    ('보기3그림', 12, IMG_FILL),
    ('보기4', 28, CONTENT_FILL),
    ('보기4그림', 12, IMG_FILL),
    ('정답(1~4)', 8, ANS_FILL),
    ('해설', 55, CONTENT_FILL),
    ('해설그림', 14, IMG_FILL),
    ('오답분석', 45, CONTENT_FILL),
    ('학습포인트', 40, CONTENT_FILL),
    ('학습포인트그림', 14, IMG_FILL),
    ('이슈', 35, ISSUE_FILL),
    ('PDF페이지', 8, META_FILL),
]

def write_excel(qs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '문항등록'

    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for i, (h, w, fill) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(1, i)
        c.font = HDR_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = HDR_BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    for q in qs:
        q = apply_manual_fixes(q)
        row = [
            q['no'],
            q['subject'] or '',
            '', '',           # 토픽/서브토픽은 수동
            q['stem'],
            '',                # 조건
            q['stem_img'],
            q['opt1'], q['opt1_img'],
            q['opt2'], q['opt2_img'],
            q['opt3'], q['opt3_img'],
            q['opt4'], q['opt4_img'],
            q['answer'],
            q['explain'],
            q['explain_img'],
            q['odab'],
            q['learning_point'],
            q['lp_img'],
            q['issues'],
            q['page_start'],
        ]
        ws.append(row)
        for c in range(1, ws.max_column + 1):
            ws.cell(ws.max_row, c).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[ws.max_row].height = 90

    # ---------- 이슈 시트 ----------
    issue_ws = wb.create_sheet('이슈리포트')
    issue_ws.append(['번호', '과목ID', '이슈', '발문 미리보기'])
    for c in range(1, 5):
        cell = issue_ws.cell(1, c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    issue_ws.column_dimensions['A'].width = 8
    issue_ws.column_dimensions['B'].width = 12
    issue_ws.column_dimensions['C'].width = 50
    issue_ws.column_dimensions['D'].width = 60
    issue_ws.row_dimensions[1].height = 24
    issue_ws.freeze_panes = 'A2'

    for q in qs:
        if not q['issues']:
            continue
        issue_ws.append([
            q['no'],
            q['subject'] or '',
            q['issues'],
            (q['stem'] or '')[:80]
        ])
        for c in range(1, 5):
            issue_ws.cell(issue_ws.max_row, c).alignment = Alignment(wrap_text=True, vertical='top')

    # ---------- 변환 통계 시트 ----------
    stat_ws = wb.create_sheet('변환통계', 0)  # 첫 시트로
    stat_ws.append(['항목', '값'])
    stat_ws['A1'].font = HDR_FONT
    stat_ws['B1'].font = HDR_FONT
    stat_ws['A1'].fill = HDR_FILL
    stat_ws['B1'].fill = HDR_FILL
    total = len(qs)
    with_issue = sum(1 for q in qs if q['issues'])
    no_answer = sum(1 for q in qs if not q['answer'])
    empty_opt = sum(1 for q in qs if not all([q['opt1'], q['opt2'], q['opt3'], q['opt4']]))
    lp_count = sum(1 for q in qs if q.get('learning_point'))
    lp_auto_count = sum(1 for q in qs if q.get('lp_auto'))
    manual_count = sum(1 for q in qs if q.get('manual_applied'))
    by_subj = {}
    for q in qs:
        by_subj[q['subject'] or '?'] = by_subj.get(q['subject'] or '?', 0) + 1

    rows = [
        ('총 문항 수', total),
        ('이슈 있는 문항', with_issue),
        ('정답 추출 실패', no_answer),
        ('보기 중 비어있는 항목 있음', empty_opt),
        ('자동 정상 변환 (이슈 없음)', total - with_issue),
        ('', ''),
        ('학습포인트 보유 문항', lp_count),
        ('  - 자동 분리', lp_auto_count),
        ('  - 수동 보완 포함', manual_count),
        ('', ''),
        ('수동 보완 적용 문항', manual_count),
        ('', ''),
        ('과목별 분포', ''),
    ]
    for s, n in by_subj.items():
        rows.append((f'  {s}', n))

    for r in rows:
        stat_ws.append(r)
    stat_ws.column_dimensions['A'].width = 30
    stat_ws.column_dimensions['B'].width = 12

    wb.save(OUT)
    print(f'WROTE: {OUT}')
    print(f'  총 {total}문항, 이슈 {with_issue}건, 정상 변환 {total - with_issue}건')
    print(f'  과목별:', by_subj)

# ---------- main ----------
if __name__ == '__main__':
    pages = load_pages()
    qs = parse_questions(pages)
    print(f'parsed {len(qs)} questions')
    structured = [structure_question(q) for q in qs]
    write_excel(structured)
