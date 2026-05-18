import openpyxl
import glob
import sys

# stdout utf-8
sys.stdout.reconfigure(encoding='utf-8')

paths = glob.glob('data/templates/*.xlsx')
for p in paths:
    if '.bak' in p:
        continue
    print(f'\n########## FILE: {p} ##########')
    wb = openpyxl.load_workbook(p, data_only=False)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f'\n=== Sheet: {sn} (rows={ws.max_row}, cols={ws.max_column}) ===')
        max_rows = min(4, ws.max_row)
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), 1):
            print(f'  R{i}:', row)
