"""문항 양방향 변환 라이브러리.

- 등록 페이지 JSON ↔ 엑셀 v3 행 ↔ 블록 배열

해설/학습포인트 셀 직렬화 규칙 (HTML-native):
  - 셀에 HTML을 그대로 작성한다.
  - 그림은 본문 어디에나 `<img src="파일명">` 으로 삽입 — 여러 개·텍스트와
    교차 가능 (해설↔그림↔해설↔그림 …).
  - 파서가 `<img>` 경계로 text/image 블록을 자동 분할한다.
  - 구 표기 `[IMG: 파일명]` 도 계속 인식한다 (하위 호환).
  - (오답분석은 블록이 아니라 raw HTML 그대로 — <img> 인라인 렌더됨.)
"""
from __future__ import annotations
import re, json
from typing import Any

# ============== 블록 ↔ 셀 ==============
# 그림 토큰: HTML <img …> 태그 또는 레거시 [IMG: 파일명]
_IMG_TOKEN_RE = re.compile(r'<img\b[^>]*>|\[IMG:\s*(.+?)\s*\]', re.I | re.S)
_SRC_RE = re.compile(
    r'src\s*=\s*(?:"([^"]*)"|\'([^\']*)\'|([^\s>]+))', re.I
)


def _img_value(token: str, legacy_group: str | None) -> str:
    """그림 토큰에서 파일명/마커 추출."""
    if legacy_group is not None:          # [IMG: …]
        return legacy_group.strip()
    m = _SRC_RE.search(token)             # <img src="…">
    if m:
        return (m.group(1) or m.group(2) or m.group(3) or '').strip()
    return ''


def blocks_to_cell(blocks: list[dict]) -> str:
    """블록 배열 → 엑셀 셀 문자열 (그림은 <img src="…">)."""
    parts = []
    for b in blocks or []:
        t = b.get('type')
        v = (b.get('value') or '').rstrip()
        if not v:
            continue
        if t == 'text':
            parts.append(v)
        elif t == 'image':
            parts.append(f'<img src="{v}">')
    return '\n\n'.join(parts)


def cell_to_blocks(cell: str) -> list[dict]:
    """엑셀 셀 문자열 → 블록 배열.
    `<img …>` 태그(또는 레거시 `[IMG: …]`)를 기준으로 앞뒤 HTML을
    text 블록, 그림 토큰을 image 블록으로 분할한다."""
    if not cell or not cell.strip():
        return []
    blocks: list[dict] = []
    pos = 0
    for m in _IMG_TOKEN_RE.finditer(cell):
        head = cell[pos:m.start()].strip()
        if head:
            blocks.append({'type': 'text', 'value': head})
        val = _img_value(m.group(0), m.group(1))
        if val:
            blocks.append({'type': 'image', 'value': val})
        pos = m.end()
    tail = cell[pos:].strip()
    if tail:
        blocks.append({'type': 'text', 'value': tail})
    return blocks


# ============== 엑셀 행(dict) ↔ 등록 페이지 JSON ==============
# 엑셀 컬럼명과 JSON 키가 거의 일치하지만, 블록 시퀀스 두 컬럼만 변환 필요.

BLOCK_COLS = ('해설', '학습포인트')   # 엑셀 컬럼명
BLOCK_JSON = ('해설블록', '학습포인트블록')  # JSON 키


def row_to_json(row: dict[str, Any]) -> dict:
    """엑셀 행(dict) → 등록 페이지 JSON 형식.
    - 해설 / 학습포인트 컬럼은 cell_to_blocks 로 배열화 (+ id 부여)"""
    out: dict[str, Any] = {}
    for k, v in row.items():
        if k == '해설':
            out['해설블록'] = _add_ids(cell_to_blocks(str(v or '')))
        elif k == '학습포인트':
            out['학습포인트블록'] = _add_ids(cell_to_blocks(str(v or '')))
        else:
            out[k] = v
    return out


def json_to_row(j: dict[str, Any]) -> dict:
    """등록 페이지 JSON → 엑셀 행(dict).
    - 해설블록 / 학습포인트블록 → blocks_to_cell 로 직렬화"""
    out: dict[str, Any] = {}
    for k, v in j.items():
        if k == '해설블록':
            out['해설'] = blocks_to_cell(v or [])
        elif k == '학습포인트블록':
            out['학습포인트'] = blocks_to_cell(v or [])
        else:
            out[k] = v
    return out


def _add_ids(blocks: list[dict]) -> list[dict]:
    """등록 페이지가 사용하는 id 필드 채워주기 (없어도 동작은 함)."""
    import uuid
    return [{**b, 'id': b.get('id') or uuid.uuid4().hex[:12]} for b in blocks]


# ============== 엑셀 파일 → JSON 리스트 ==============
# v3 양식 1행은 그룹 헤더(병합)라 "출처/코드/강의/..." 같은 한글 라벨이 옴.
# 그 라벨이 컬럼명으로 다시 등장하는지 보고 v3(2단)인지 v2(1단)인지 자동 판별.
_GROUP_HEADER_LABELS = {'번호', '출처', '코드', '강의', '분류', '속성', '발문',
                         '보기', '정답', '해설', '오답분석', '학습POINT'}

def xlsx_to_json_list(path: str) -> list[dict]:
    """엑셀 v2/v3 양식 → 문항 JSON 배열. 1행이 그룹 헤더이면 2행을 헤더로."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if '문항등록' not in wb.sheetnames:
        raise ValueError(f'시트 "문항등록" 없음: {wb.sheetnames}')
    ws = wb['문항등록']

    # v3 판별: 2행에 컬럼 헤더('문항코드'/'발문')가 있으면 1행은 그룹 헤더(2단).
    # ('번호'처럼 그룹 라벨과 컬럼명이 같을 수 있어 1행만으로는 판별 불가)
    def _row_vals(r):
        return {str(ws.cell(r, c).value or '').strip()
                for c in range(1, ws.max_column + 1)}
    row2 = _row_vals(2)
    is_v3 = bool({'문항코드', '발문'} & row2)
    header_row = 2 if is_v3 else 1
    data_start = header_row + 1

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(data_start, ws.max_row + 1):
        row = {}
        empty = True
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if v is not None and v != '':
                empty = False
            row[h] = v
        if not empty:
            out.append(row_to_json(row))
    return out


# ============== JSON 리스트 → 엑셀 파일 ==============
def json_list_to_xlsx(items: list[dict], path: str):
    """JSON 배열 → 엑셀 v3 양식. (build_template_v3.py와 같은 컬럼 구조 가정)"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    # 기본 헤더 (build_template_v3.py와 동일 순서)
    headers = [
        '번호',
        '과정', '연도', '회차', '사용교재', '교재구분',
        '문항코드',
        '강의주소',
        '과목ID', '챕터', '대유형', '중유형', '내용',
        '빈출도', '난이도', '문제유형', '변형이력', '비고',
        '발문', '조건', '발문그림',
        '보기1', '보기1그림', '보기2', '보기2그림',
        '보기3', '보기3그림', '보기4', '보기4그림',
        '정답(1~4)',
        '해설', '오답분석', '학습포인트',
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '문항등록'
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
    for j in items:
        row_dict = json_to_row(j)
        ws.append([row_dict.get(h, '') for h in headers])
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 120
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    wb.save(path)


# ============== CLI ==============
if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print('usage:')
        print('  python question_io.py xlsx2json <input.xlsx> [output.json]')
        print('  python question_io.py json2xlsx <input.json>  [output.xlsx]')
        sys.exit(1)

    sys.stdout.reconfigure(encoding='utf-8')
    cmd, src = sys.argv[1], sys.argv[2]
    dst = sys.argv[3] if len(sys.argv) > 3 else None

    if cmd == 'xlsx2json':
        items = xlsx_to_json_list(src)
        dst = dst or src.replace('.xlsx', '.json')
        with open(dst, 'w', encoding='utf-8') as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
        print(f'WROTE {dst}  ({len(items)} 문항)')
    elif cmd == 'json2xlsx':
        with open(src, 'r', encoding='utf-8') as f:
            items = json.load(f)
        if isinstance(items, dict):
            items = [items]
        dst = dst or src.replace('.json', '.xlsx')
        json_list_to_xlsx(items, dst)
        print(f'WROTE {dst}  ({len(items)} 문항)')
    else:
        print(f'unknown command: {cmd}')
        sys.exit(1)
